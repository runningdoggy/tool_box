#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

try:
    from opencc import OpenCC
except Exception:
    OpenCC = None


@dataclass
class CalibrateResult:
    fixed_text: str
    changed: bool
    score: int
    level: str
    notes: str


REPLACEMENTS: list[tuple[re.Pattern[str], str, str]] = [
    (re.compile(r"沒關係"), "没关系", "繁简转换"),
    (re.compile(r"這個"), "这个", "繁简转换"),
    (re.compile(r"視頻"), "视频", "繁简转换"),
    (re.compile(r"學習|學"), "学习", "繁简转换"),
    (re.compile(r"賺錢"), "赚钱", "繁简转换"),
    (re.compile(r"課程"), "课程", "繁简转换"),
    (re.compile(r"免費"), "免费", "繁简转换"),
    (re.compile(r"手機"), "手机", "繁简转换"),
    (re.compile(r"時間"), "时间", "繁简转换"),
    (re.compile(r"關係"), "关系", "繁简转换"),
    (re.compile(r"國債"), "国债", "繁简转换"),
    (re.compile(r"風險"), "风险", "繁简转换"),
    (re.compile(r"點位"), "点位", "繁简转换"),
    (re.compile(r"邏輯"), "逻辑", "繁简转换"),
    (re.compile(r"適合"), "适合", "繁简转换"),
    (re.compile(r"咱們"), "咱们", "繁简转换"),
    (re.compile(r"別"), "别", "繁简转换"),
    (re.compile(r"盯盤"), "盯盘", "繁简转换"),
    (re.compile(r"聽"), "听", "繁简转换"),
    (re.compile(r"賣"), "卖", "繁简转换"),
    (re.compile(r"買"), "买", "繁简转换"),
    (re.compile(r"歲"), "岁", "繁简转换"),
    (re.compile(r"塊"), "块", "繁简转换"),
    (re.compile(r"漲"), "涨", "繁简转换"),
    (re.compile(r"裡"), "里", "繁简转换"),
    (re.compile(r"[eE][tT][fF]"), "ETF", "ETF大小写统一"),
    (re.compile(r"(?i)\b[1Il]\s*T\s*F\b"), "ETF", "ETF误识别修正"),
    (re.compile(r"一\s*T\s*F"), "ETF", "ETF误识别修正"),
    (re.compile(r"(?i)\bE\s*T\s*F\b"), "ETF", "ETF空格归一"),
    (re.compile(r"ETF\s*ETF"), "ETF", "ETF重复去重"),
    (re.compile(r"[一1I]\s*天\s*[fF夫副份赋富否不布]"), "ETF", "ETF近音词修正"),
    (re.compile(r"[一1I]\s*[帖贴貼铁鐵條条]\s*[fF夫粉]"), "ETF", "ETF近音词修正"),
    (re.compile(r"[一1I]\s*贴\s*富"), "ETF", "ETF近音词修正"),
    (re.compile(r"[一1I]\s*带\s*[fF]"), "ETF", "ETF近音词修正"),
    (re.compile(r"一天\s*[fF夫副份赋富否不布]"), "ETF", "ETF近音词修正"),
    (re.compile(r"一体[儿兒]子|一體兒子"), "ETF", "ETF近音词修正"),
    (re.compile(r"一条比赛|一條比賽"), "ETF理财", "ETF短语修正"),
    (
        re.compile(
            r"一?天[赋富复復負负福幅扶府否不布副夫份](?=(给|就|是|都|很|能|来|去|做|买|卖|投资|指数|基金|理财|课程|工具|买卖|策略|技巧|资产|交易|入门|启蒙|起门|起盟|礼财|离财|里财|理裁|理才))"
        ),
        "ETF",
        "ETF近音词修正",
    ),
    (re.compile(r"理材|礼材|离才|礼财|李财|里财|离财|理裁|理才|里裁|里才|底裁|禮財|禮材"), "理财", "理财词修正"),
    (re.compile(r"理[财彩餐]客"), "理财课", "课程词修正"),
    (re.compile(r"理[财彩餐]课"), "理财课", "课程词修正"),
    (re.compile(r"ETF餐客"), "ETF理财课", "课程词修正"),
    (re.compile(r"练坚|练接|鏈接|鏈坚|连结|連接|鏈結"), "链接", "链接词修正"),
    (re.compile(r"听可|停可|聽可|聽課"), "听课", "听课词修正"),
    (re.compile(r"酒天|久天|九天"), "9天", "数字词修正"),
    (re.compile(r"丁盘"), "盯盘", "术语修正"),
    (re.compile(r"启早|七早"), "起早", "常见口语修正"),
    (re.compile(r"探黑"), "贪黑", "常见口语修正"),
    (re.compile(r"试做成本"), "试错成本", "理财词修正"),
    (re.compile(r"灵回去|令回去"), "领回去", "词语修正"),
    (re.compile(r"门盘是中"), "门槛适中", "语义修正"),
    (re.compile(r"风格文件"), "风格稳健", "语义修正"),
    (re.compile(r"不动可控"), "波动可控", "语义修正"),
    (re.compile(r"一男子优质资产|一栏子优质资产|一蓝子优质资产"), "一篮子优质资产", "术语修正"),
    (re.compile(r"文件[吃持]有方法"), "稳健持有方法", "语义修正"),
    (re.compile(r"存因行立西|存因行利西|存银行立息|存因行里西"), "存银行利息", "术语修正"),
    (re.compile(r"潮谷|超股|操股|朝鲍|潮谱"), "炒股", "术语修正"),
    (re.compile(r"朝骨票|跑骨票"), "炒股票", "术语修正"),
    (re.compile(r"产米游言|产品游言|产米油言|柴米游言"), "柴米油盐", "词语修正"),
    (re.compile(r"安门[门闻文]?想[清亲幸]服|安门门想清楚"), "安稳享清福", "短语修正"),
    (re.compile(r"文本的收益|闻本的收益|门门的收益"), "稳稳的收益", "短语修正"),
    (re.compile(r"宝险|保鞋|保鲜"), "保险", "术语修正"),
    (re.compile(r"本经"), "本金", "术语修正"),
    (re.compile(r"门卡"), "门槛", "术语修正"),
    (re.compile(r"采坑|财坑|彩坑"), "踩坑", "术语修正"),
    (re.compile(r"(?<=了解一下)一天(?=就是)"), "ETF", "ETF语义修正"),
    (re.compile(r"两倍的9天"), "免费的9天", "短语修正"),
]


TRADITIONAL_CHAR_MAP = str.maketrans(
    {
        "萬": "万",
        "與": "与",
        "專": "专",
        "業": "业",
        "東": "东",
        "兩": "两",
        "個": "个",
        "中": "中",
        "為": "为",
        "麼": "么",
        "義": "义",
        "之": "之",
        "習": "习",
        "買": "买",
        "亂": "乱",
        "乾": "干",
        "了": "了",
        "爭": "争",
        "事": "事",
        "亞": "亚",
        "產": "产",
        "們": "们",
        "優": "优",
        "會": "会",
        "傷": "伤",
        "價": "价",
        "體": "体",
        "來": "来",
        "個": "个",
        "倉": "仓",
        "倫": "伦",
        "傳": "传",
        "傷": "伤",
        "債": "债",
        "儲": "储",
        "兒": "儿",
        "兩": "两",
        "關": "关",
        "內": "内",
        "寫": "写",
        "劃": "划",
        "劑": "剂",
        "動": "动",
        "勵": "励",
        "勻": "匀",
        "勢": "势",
        "區": "区",
        "醫": "医",
        "華": "华",
        "單": "单",
        "嗎": "吗",
        "啟": "启",
        "問": "问",
        "國": "国",
        "圖": "图",
        "團": "团",
        "壓": "压",
        "壞": "坏",
        "壽": "寿",
        "夠": "够",
        "夢": "梦",
        "夥": "伙",
        "學": "学",
        "實": "实",
        "對": "对",
        "導": "导",
        "將": "将",
        "專": "专",
        "尋": "寻",
        "層": "层",
        "屬": "属",
        "歲": "岁",
        "島": "岛",
        "師": "师",
        "帳": "账",
        "帶": "带",
        "幣": "币",
        "幫": "帮",
        "庫": "库",
        "應": "应",
        "廠": "厂",
        "廣": "广",
        "廳": "厅",
        "彈": "弹",
        "強": "强",
        "當": "当",
        "彙": "汇",
        "後": "后",
        "徑": "径",
        "從": "从",
        "復": "复",
        "德": "德",
        "憂": "忧",
        "應": "应",
        "戲": "戏",
        "戶": "户",
        "手": "手",
        "拋": "抛",
        "擔": "担",
        "據": "据",
        "數": "数",
        "敵": "敌",
        "斷": "断",
        "新": "新",
        "時": "时",
        "晉": "晋",
        "暫": "暂",
        "會": "会",
        "曉": "晓",
        "書": "书",
        "會": "会",
        "術": "术",
        "條": "条",
        "來": "来",
        "東": "东",
        "極": "极",
        "標": "标",
        "樣": "样",
        "機": "机",
        "權": "权",
        "歡": "欢",
        "歐": "欧",
        "歲": "岁",
        "歷": "历",
        "歸": "归",
        "殘": "残",
        "氣": "气",
        "沒": "没",
        "灣": "湾",
        "滿": "满",
        "滅": "灭",
        "漲": "涨",
        "潤": "润",
        "澤": "泽",
        "無": "无",
        "點": "点",
        "為": "为",
        "熱": "热",
        "營": "营",
        "獎": "奖",
        "環": "环",
        "現": "现",
        "產": "产",
        "畫": "画",
        "當": "当",
        "發": "发",
        "眾": "众",
        "盤": "盘",
        "盯": "盯",
        "監": "监",
        "瞭": "了",
        "礎": "础",
        "禮": "礼",
        "種": "种",
        "穩": "稳",
        "窩": "窝",
        "簡": "简",
        "築": "筑",
        "類": "类",
        "糾": "纠",
        "經": "经",
        "綱": "纲",
        "網": "网",
        "總": "总",
        "線": "线",
        "績": "绩",
        "續": "续",
        "聽": "听",
        "職": "职",
        "聯": "联",
        "聲": "声",
        "聰": "聪",
        "腦": "脑",
        "臺": "台",
        "與": "与",
        "興": "兴",
        "舉": "举",
        "舊": "旧",
        "艙": "舱",
        "節": "节",
        "藝": "艺",
        "蘭": "兰",
        "虧": "亏",
        "號": "号",
        "處": "处",
        "術": "术",
        "視": "视",
        "覺": "觉",
        "覽": "览",
        "觀": "观",
        "計": "计",
        "訓": "训",
        "訊": "讯",
        "記": "记",
        "設": "设",
        "證": "证",
        "評": "评",
        "詐": "诈",
        "試": "试",
        "話": "话",
        "該": "该",
        "詳": "详",
        "語": "语",
        "說": "说",
        "課": "课",
        "請": "请",
        "讀": "读",
        "變": "变",
        "讓": "让",
        "議": "议",
        "負": "负",
        "財": "财",
        "貝": "贝",
        "貴": "贵",
        "費": "费",
        "資": "资",
        "賣": "卖",
        "賺": "赚",
        "趕": "赶",
        "車": "车",
        "較": "较",
        "輕": "轻",
        "辦": "办",
        "這": "这",
        "還": "还",
        "邏": "逻",
        "邊": "边",
        "醫": "医",
        "釋": "释",
        "鐘": "钟",
        "錢": "钱",
        "錄": "录",
        "錯": "错",
        "門": "门",
        "關": "关",
        "開": "开",
        "闆": "板",
        "險": "险",
        "雙": "双",
        "雜": "杂",
        "靈": "灵",
        "靜": "静",
        "風": "风",
        "飛": "飞",
        "養": "养",
        "騙": "骗",
        "驗": "验",
        "點": "点",
    }
)


SUSPICIOUS_PATTERNS: list[tuple[re.Pattern[str], str, int]] = [
    (re.compile(r"(?<!ETF)[A-Za-z]{2,}"), "存在异常英文片段", 8),
    (re.compile(r"[^\u4e00-\u9fffA-Za-z0-9，。！？、；：“”‘’（）《》【】\-\s]"), "存在异常字符", 6),
    (re.compile(r"[一1I][天帖贴鐵铁條条体體][fF夫副份赋富否不布粉]"), "疑似ETF词仍未清洗", 8),
]

OPENCC_CONVERTER = OpenCC("t2s") if OpenCC is not None else None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="对已有素材报表中的脚本列进行二次校准。")
    parser.add_argument("--input", required=True, help="输入 xlsx 路径")
    parser.add_argument("--output", default="", help="输出 xlsx 路径，默认 input_stem + _校准.xlsx")
    parser.add_argument("--sheet", default="", help="sheet 名，不填则使用第一个")
    parser.add_argument("--script-col", default="脚本信息", help="待校准列名")
    parser.add_argument("--output-col", default="脚本信息_校准", help="输出校准文本列名")
    parser.add_argument("--score-col", default="脚本质量评分", help="质量评分列名")
    parser.add_argument("--level-col", default="脚本质量等级", help="质量等级列名")
    parser.add_argument("--note-col", default="校准说明", help="校准说明列名")
    return parser.parse_args()


def normalize_text(text: str) -> str:
    fixed = unicodedata.normalize("NFKC", text or "")
    if OPENCC_CONVERTER is not None:
        fixed = OPENCC_CONVERTER.convert(fixed)
    fixed = fixed.translate(TRADITIONAL_CHAR_MAP)
    fixed = fixed.replace("\u3000", " ")
    fixed = re.sub(r"\s+", " ", fixed).strip()
    fixed = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])", "", fixed)
    fixed = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=ETF)", "", fixed)
    fixed = re.sub(r"(?<=ETF)\s+(?=[\u4e00-\u9fff])", "", fixed)
    return fixed


def run_replacements(text: str) -> tuple[str, list[str]]:
    current = normalize_text(text)
    notes: list[str] = []
    for pattern, target, note in REPLACEMENTS:
        updated = pattern.sub(target, current)
        if updated != current:
            notes.append(note)
            current = updated
    updated = re.sub(r"(?i)\b[A-Za-z]*\s*E\s*T\s*F[A-Za-z]*\b", "ETF", current)
    if updated != current:
        notes.append("ETF字母噪声清理")
        current = updated
    updated = re.sub(r"ETF[A-Za-z]+", "ETF", current)
    if updated != current:
        notes.append("ETF后缀噪声清理")
        current = updated
    updated = re.sub(r"(?<!ETF)\b[A-Za-z]{2,}\b", "", current)
    if updated != current:
        notes.append("英文噪声清理")
        current = updated
    current = re.sub(r"\s+", " ", current).strip()
    current = re.sub(r"[，,]{2,}", "，", current)
    current = re.sub(r"[。\.]{2,}", "。", current)
    current = re.sub(r"[!！]{2,}", "！", current)
    current = re.sub(r"[?？]{2,}", "？", current)
    current = current.strip("，。；,; ")
    return current, list(dict.fromkeys(notes))


def quality_check(text: str) -> tuple[int, str, list[str]]:
    if not text:
        return 20, "低", ["空文本"]
    score = 100
    notes: list[str] = []
    if len(text) < 30:
        score -= 22
        notes.append("文本较短")
    if "ETF" not in text:
        score -= 16
        notes.append("缺少ETF关键词")
    for pattern, note, penalty in SUSPICIOUS_PATTERNS:
        if pattern.search(text):
            score -= penalty
            notes.append(note)
    score = max(0, min(100, score))
    if score >= 85:
        level = "高"
    elif score >= 70:
        level = "中"
    else:
        level = "低"
    return score, level, notes


def calibrate_text(raw: str) -> CalibrateResult:
    fixed, fix_notes = run_replacements(raw)
    score, level, check_notes = quality_check(fixed)
    notes = "；".join(list(dict.fromkeys(fix_notes + check_notes)))
    return CalibrateResult(
        fixed_text=fixed,
        changed=(fixed != (raw or "").strip()),
        score=score,
        level=level,
        notes=notes,
    )


def ensure_column(headers: list[str], wanted_name: str) -> int:
    if wanted_name in headers:
        return headers.index(wanted_name) + 1
    headers.append(wanted_name)
    return len(headers)


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"输入文件不存在: {input_path}")
        return 1
    output_path = Path(args.output) if args.output else input_path.with_name(f"{input_path.stem}_校准.xlsx")
    workbook = load_workbook(input_path)
    sheet = workbook[args.sheet] if args.sheet else workbook[workbook.sheetnames[0]]
    max_col = sheet.max_column
    headers = [str(sheet.cell(1, col).value or "").strip() for col in range(1, max_col + 1)]
    if args.script_col not in headers:
        print(f"未找到脚本列: {args.script_col}")
        workbook.close()
        return 1
    script_col_idx = headers.index(args.script_col) + 1
    output_col_idx = ensure_column(headers, args.output_col)
    score_col_idx = ensure_column(headers, args.score_col)
    level_col_idx = ensure_column(headers, args.level_col)
    note_col_idx = ensure_column(headers, args.note_col)
    for col_idx, title in enumerate(headers, start=1):
        sheet.cell(1, col_idx).value = title

    total = 0
    changed = 0
    high = 0
    mid = 0
    low = 0
    for row in range(2, sheet.max_row + 1):
        raw_script = str(sheet.cell(row, script_col_idx).value or "").strip()
        if not raw_script:
            continue
        total += 1
        result = calibrate_text(raw_script)
        if result.changed:
            changed += 1
        if result.level == "高":
            high += 1
        elif result.level == "中":
            mid += 1
        else:
            low += 1
        sheet.cell(row, output_col_idx).value = result.fixed_text
        sheet.cell(row, score_col_idx).value = result.score
        sheet.cell(row, level_col_idx).value = result.level
        sheet.cell(row, note_col_idx).value = result.notes

    workbook.save(output_path)
    workbook.close()
    print(f"处理完成: {total} 行 | 有改动: {changed}")
    print(f"质量等级: 高={high} 中={mid} 低={low}")
    print(f"输出文件: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
