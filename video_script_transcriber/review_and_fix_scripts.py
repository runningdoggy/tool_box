#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import os
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from huggingface_hub import InferenceClient
from openpyxl import Workbook, load_workbook


@dataclass
class ReviewRow:
    url: str
    raw_script: str
    fixed_script: str
    changed: bool
    llm_used: bool
    score: int
    level: str
    notes: str


PHRASE_REPLACEMENTS: list[tuple[re.Pattern[str], str, str]] = [
    (re.compile(r"1\s*\.?\s*[fF]"), "ETF", "ETF写法统一"),
    (re.compile(r"(?i)[1Il]\s*T\s*F"), "ETF", "ETF误识别修正"),
    (re.compile(r"一\s*T\s*F"), "ETF", "ETF误识别修正"),
    (re.compile(r"(?i)\bE\s*T\s*F\b"), "ETF", "ETF空格归一"),
    (
        re.compile(r"一[天条條帖贴貼]\s*[复復负負幅扶服福府]"),
        "ETF",
        "ETF近音词修正",
    ),
    (
        re.compile(
            r"一?天[赋富复復負负福幅扶府](?=(给|就|是|都|很|能|来|去|做|买|卖|投资|指数|基金|理财|里财|离财|理裁|课程|工具|买卖|策略|技巧|资产|交易|入门|起蒙|起门|起盟))"
        ),
        "ETF",
        "ETF近音词修正",
    ),
    (re.compile(r"[eE][tT][fF]"), "ETF", "ETF大小写统一"),
    (re.compile(r"ETF\s*ETF"), "ETF", "重复ETF去重"),
    (re.compile(r"(?<=\d)\s*T\s*ETF"), "个ETF", "数量词修正"),
    (re.compile(r"国債"), "国债", "繁简字修正"),
    (re.compile(r"理材|礼材|离才|礼财|李财|里财|离财|理裁|理才"), "理财", "理财词修正"),
    (re.compile(r"练坚|练接|鏈接|鏈坚|连结|連接|鏈結"), "链接", "链接词修正"),
    (re.compile(r"听可|停可|聽可"), "听课", "听课词修正"),
]


SUSPICIOUS_PATTERNS: list[tuple[re.Pattern[str], str, int]] = [
    (re.compile(r"[A-DF-Za-df-z]{3,}"), "存在异常英文片段", 8),
    (re.compile(r"\d+\.[A-Za-z]"), "存在疑似识别噪声(如 1.F)", 5),
    (re.compile(r"[^\u4e00-\u9fffA-Za-z0-9，。！？、；：“”‘’（）\-\s]"), "存在异常字符", 6),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="对转写脚本做逻辑修正与质检打标。")
    parser.add_argument("--input", required=True, help="输入汇总 Excel（默认列：链接、脚本）")
    parser.add_argument("--output-dir", default="", help="输出目录（默认与输入同目录）")
    parser.add_argument(
        "--output-prefix",
        default="scripts_corrected",
        help="输出文件名前缀（默认：scripts_corrected）",
    )
    parser.add_argument(
        "--llm-backend",
        choices=("none", "hf"),
        default="none",
        help="是否启用开源大模型校准（none/hf）",
    )
    parser.add_argument(
        "--llm-model",
        default="Qwen/Qwen2.5-7B-Instruct",
        help="HF模型名（默认：Qwen/Qwen2.5-7B-Instruct）",
    )
    parser.add_argument(
        "--llm-scope",
        choices=("low", "suspicious", "all"),
        default="low",
        help="哪些行走LLM校准：low/suspicious/all",
    )
    parser.add_argument(
        "--llm-max-rows",
        type=int,
        default=80,
        help="最多送入LLM的行数（默认80）",
    )
    parser.add_argument(
        "--hf-token",
        default=os.getenv("HF_TOKEN", ""),
        help="Hugging Face Token（默认读取环境变量 HF_TOKEN）",
    )
    return parser.parse_args()


def normalize_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKC", text or "")
    normalized = re.sub(r"\s+", " ", normalized).strip()
    normalized = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])", "", normalized)
    normalized = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=ETF)", "", normalized)
    normalized = re.sub(r"(?<=ETF)\s+(?=[\u4e00-\u9fff])", "", normalized)
    return normalized


def apply_corrections(text: str) -> tuple[str, list[str]]:
    current = normalize_text(text)
    notes: list[str] = []
    for pattern, target, note in PHRASE_REPLACEMENTS:
        updated = pattern.sub(target, current)
        if updated != current:
            notes.append(note)
            current = updated
    current = re.sub(r"\s+", " ", current).strip()
    return current, notes


def quality_check(text: str) -> tuple[int, str, list[str]]:
    score = 100
    notes: list[str] = []
    if not text:
        return 30, "低", ["空文本"]
    if len(text) < 30:
        score -= 25
        notes.append("文本较短")
    if "ETF" not in text and ("理财" in text or "国债" in text):
        score -= 8
        notes.append("疑似ETF词缺失")
    for pattern, note, penalty in SUSPICIOUS_PATTERNS:
        if pattern.search(text):
            score -= penalty
            notes.append(note)
    score = max(0, min(100, score))
    if score >= 85:
        level = "高"
    elif score >= 70:
        level = "中"
    else:
        level = "低"
    return score, level, notes


def is_suspicious_text(text: str) -> bool:
    lowered = text.lower()
    if any(token in lowered for token in ("1tf", "itf", "ltf", "et f")):
        return True
    if re.search(r"[一天条條帖贴貼][赋富复復負负福幅扶府]", text):
        return True
    return any(pattern.search(text) for pattern, _, _ in SUSPICIOUS_PATTERNS)


def should_run_llm(scope: str, level: str, text: str) -> bool:
    if scope == "all":
        return True
    if scope == "low":
        return level == "低"
    return is_suspicious_text(text)


def llm_correct_hf(
    client: InferenceClient,
    model: str,
    text: str,
) -> tuple[str, str | None]:
    try:
        result = client.chat_completion(
            model=model,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "你是中文ASR脚本校对器。任务："
                        "1) 只修正常见错别字/同音误识别；"
                        "2) 将 1TF/ITF/一TF/一天赋(在投资语境) 统一成 ETF；"
                        "3) 不要扩写、不删减事实、不加解释；"
                        "4) 输出仅保留校准后的正文。"
                    ),
                },
                {
                    "role": "user",
                    "content": text,
                },
            ],
            temperature=0.1,
            max_tokens=2048,
        )
        content = (result.choices[0].message.content or "").strip()
        content = content.strip("“”\"'")
        if not content:
            return text, "LLM空输出，已回退"
        return content, None
    except Exception as error:
        return text, f"LLM失败：{error}"


def read_input_rows(input_path: Path) -> list[tuple[str, str]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    _ = next(iterator, None)  # header
    rows: list[tuple[str, str]] = []
    for row in iterator:
        if not row:
            continue
        url = str(row[0] or "").strip()
        script = str(row[1] or "").strip()
        if not url:
            continue
        rows.append((url, script))
    workbook.close()
    return rows


def build_review_rows(raw_rows: list[tuple[str, str]]) -> list[ReviewRow]:
    return build_review_rows_with_llm(
        raw_rows=raw_rows,
        llm_backend="none",
        llm_model="",
        llm_scope="low",
        llm_max_rows=0,
        hf_token="",
    )


def build_review_rows_with_llm(
    raw_rows: list[tuple[str, str]],
    llm_backend: str,
    llm_model: str,
    llm_scope: str,
    llm_max_rows: int,
    hf_token: str,
) -> list[ReviewRow]:
    reviewed: list[ReviewRow] = []
    llm_calls = 0
    client: InferenceClient | None = None

    if llm_backend == "hf":
        if not hf_token:
            print("警告：未提供 HF_TOKEN，已自动关闭 LLM 校准。")
            llm_backend = "none"
        else:
            client = InferenceClient(api_key=hf_token)

    for url, raw_script in raw_rows:
        fixed_script, fix_notes = apply_corrections(raw_script)
        llm_notes: list[str] = []
        llm_used = False

        pre_score, pre_level, pre_check_notes = quality_check(fixed_script)
        if (
            llm_backend == "hf"
            and client is not None
            and llm_calls < llm_max_rows
            and should_run_llm(llm_scope, pre_level, fixed_script)
        ):
            llm_used = True
            llm_calls += 1
            llm_output, llm_error = llm_correct_hf(
                client=client,
                model=llm_model,
                text=fixed_script,
            )
            fixed_script = normalize_text(llm_output)
            fixed_script, post_fix_notes = apply_corrections(fixed_script)
            fix_notes.extend(post_fix_notes)
            if llm_error:
                llm_notes.append(llm_error)
            else:
                llm_notes.append("LLM校准")

        score, level, check_notes = quality_check(fixed_script)
        all_notes = list(dict.fromkeys(fix_notes + pre_check_notes + check_notes + llm_notes))
        reviewed.append(
            ReviewRow(
                url=url,
                raw_script=raw_script,
                fixed_script=fixed_script,
                changed=fixed_script != raw_script,
                llm_used=llm_used,
                score=score,
                level=level,
                notes="；".join(all_notes),
            )
        )
    return reviewed


def write_outputs(output_dir: Path, prefix: str, rows: list[ReviewRow]) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    excel_path = output_dir / f"{prefix}.xlsx"
    txt_path = output_dir / f"{prefix}.txt"
    csv_path = output_dir / f"{prefix}_review.csv"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "scripts_corrected"
    sheet.append(
        ["链接", "原脚本", "修正脚本", "是否修正", "是否LLM", "质量评分", "质量等级", "质检说明"]
    )
    for row in rows:
        sheet.append(
            [
                row.url,
                row.raw_script,
                row.fixed_script,
                "是" if row.changed else "否",
                "是" if row.llm_used else "否",
                row.score,
                row.level,
                row.notes,
            ]
        )
    workbook.save(excel_path)

    with txt_path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(f"{row.url}\t{row.fixed_script}\n")

    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["url", "changed", "llm_used", "score", "level", "notes"])
        for row in rows:
            writer.writerow([row.url, int(row.changed), int(row.llm_used), row.score, row.level, row.notes])

    return {"excel": excel_path, "txt": txt_path, "csv": csv_path}


def print_summary(rows: list[ReviewRow], paths: dict[str, Path]) -> None:
    total = len(rows)
    changed = sum(1 for row in rows if row.changed)
    llm_used = sum(1 for row in rows if row.llm_used)
    high = sum(1 for row in rows if row.level == "高")
    mid = sum(1 for row in rows if row.level == "中")
    low = sum(1 for row in rows if row.level == "低")
    print(f"总数: {total} | 已修正: {changed} | LLM校准: {llm_used}")
    print(f"质量等级: 高={high} 中={mid} 低={low}")
    print(f"修正版Excel: {paths['excel']}")
    print(f"修正版TXT: {paths['txt']}")
    print(f"质检明细CSV: {paths['csv']}")


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"输入文件不存在: {input_path}")
        return 1
    output_dir = Path(args.output_dir) if args.output_dir else input_path.parent
    raw_rows = read_input_rows(input_path)
    if not raw_rows:
        print("输入文件没有可处理数据。")
        return 1
    reviewed_rows = build_review_rows_with_llm(
        raw_rows=raw_rows,
        llm_backend=args.llm_backend,
        llm_model=args.llm_model,
        llm_scope=args.llm_scope,
        llm_max_rows=args.llm_max_rows,
        hf_token=args.hf_token,
    )
    paths = write_outputs(output_dir, args.output_prefix, reviewed_rows)
    print_summary(reviewed_rows, paths)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
