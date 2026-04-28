#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import re
import shutil
import subprocess
import sys
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import unquote, urlparse

from faster_whisper import WhisperModel
from openpyxl import Workbook
from openpyxl import load_workbook

try:
    from tqdm import tqdm
except ImportError:  # pragma: no cover
    tqdm = None

try:
    import imageio_ffmpeg
except ImportError:  # pragma: no cover
    imageio_ffmpeg = None


URL_PATTERN = re.compile(r"https?://\S+")
PREFERRED_URL_COLUMNS = ("url", "video_url", "link", "mp4_url")


@dataclass
class TaskResult:
    index: int
    name: str
    url: str
    status: str
    transcript_text: str | None
    error: str | None
    elapsed_seconds: float | None
    detected_language: str | None


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="批量下载 MP4 并转写为脚本文本（Whisper）。"
    )
    parser.add_argument("--input", required=True, help="链接列表文件（txt/csv/tsv）")
    parser.add_argument(
        "--output-dir", default="output", help="输出目录（默认：output）"
    )
    parser.add_argument("--model", default="small", help="Whisper 模型（默认：small）")
    parser.add_argument(
        "--device",
        default="auto",
        choices=("auto", "cpu", "cuda"),
        help="运行设备（默认：auto）",
    )
    parser.add_argument(
        "--compute-type",
        default="int8",
        help="faster-whisper 计算精度（默认：int8）",
    )
    parser.add_argument(
        "--language",
        default="auto",
        help="语言代码，例如 zh/en；auto 为自动检测",
    )
    parser.add_argument("--timeout", type=int, default=90, help="读取超时（秒）")
    parser.add_argument("--keep-audio", action="store_true", help="保留中间音频文件")
    parser.add_argument(
        "--ffmpeg-bin",
        default="auto",
        help="ffmpeg 可执行路径，默认自动查找（系统 ffmpeg 或 imageio-ffmpeg）",
    )
    return parser.parse_args()


def sanitize_filename(raw_value: str) -> str:
    cleaned = re.sub(r"[^\w.-]+", "_", raw_value.strip())
    return cleaned[:120].strip("._") or "video"


def extract_url_from_row(row: dict[str, str]) -> str | None:
    lowered_row = {key.lower().strip(): (value or "").strip() for key, value in row.items()}
    for column_name in PREFERRED_URL_COLUMNS:
        candidate = lowered_row.get(column_name)
        if candidate and URL_PATTERN.search(candidate):
            return URL_PATTERN.search(candidate).group(0)
    for value in lowered_row.values():
        match = URL_PATTERN.search(value)
        if match:
            return match.group(0)
    return None


def load_urls(input_path: Path) -> list[str]:
    if not input_path.exists():
        raise FileNotFoundError(f"未找到输入文件：{input_path}")

    urls: list[str] = []
    suffix = input_path.suffix.lower()

    if suffix in {".csv", ".tsv"}:
        delimiter = "," if suffix == ".csv" else "\t"
        with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            if not reader.fieldnames:
                return urls
            for row in reader:
                url = extract_url_from_row(row)
                if url:
                    urls.append(url)
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        candidates: list[str] = []
                        if cell.value is not None:
                            candidates.append(str(cell.value).strip())
                        hyperlink = getattr(cell, "hyperlink", None)
                        if hyperlink and getattr(hyperlink, "target", None):
                            candidates.append(str(hyperlink.target).strip())
                        for candidate in candidates:
                            if not candidate:
                                continue
                            match = URL_PATTERN.search(candidate)
                            if match:
                                urls.append(match.group(0))
        finally:
            workbook.close()
    else:
        with input_path.open("r", encoding="utf-8-sig") as handle:
            for line in handle:
                text = line.strip()
                if not text or text.startswith("#"):
                    continue
                match = URL_PATTERN.search(text)
                if match:
                    urls.append(match.group(0))

    deduplicated: list[str] = []
    seen: set[str] = set()
    for url in urls:
        if url not in seen:
            deduplicated.append(url)
            seen.add(url)
    return deduplicated


def build_base_name(url: str, index: int, used_names: set[str]) -> str:
    parsed = urlparse(url)
    source_name = Path(unquote(parsed.path)).stem
    if not source_name:
        source_name = f"video_{index:04d}_{hashlib.md5(url.encode('utf-8')).hexdigest()[:10]}"
    base_name = sanitize_filename(source_name)

    unique_name = base_name
    suffix_number = 1
    while unique_name in used_names:
        unique_name = f"{base_name}_{suffix_number}"
        suffix_number += 1
    used_names.add(unique_name)
    return unique_name


def ensure_directories(output_dir: Path) -> dict[str, Path]:
    directories = {
        "audio": output_dir / "audio",
    }
    for directory in directories.values():
        directory.mkdir(parents=True, exist_ok=True)
    return directories


def resolve_ffmpeg_binary(ffmpeg_bin: str) -> str:
    if ffmpeg_bin != "auto":
        if Path(ffmpeg_bin).exists() or shutil.which(ffmpeg_bin):
            return ffmpeg_bin
        raise EnvironmentError(f"未找到 ffmpeg：{ffmpeg_bin}")

    if shutil.which("ffmpeg"):
        return "ffmpeg"
    if imageio_ffmpeg is not None:
        return imageio_ffmpeg.get_ffmpeg_exe()
    raise EnvironmentError(
        "未检测到 ffmpeg。请安装系统 ffmpeg，或执行 `pip install imageio-ffmpeg`。"
    )


def extract_audio_from_url(
    ffmpeg_bin: str, video_url: str, audio_path: Path, timeout: int
) -> None:
    timeout_us = max(1, timeout) * 1_000_000
    command = [
        ffmpeg_bin,
        "-y",
        "-rw_timeout",
        str(timeout_us),
        "-i",
        video_url,
        "-vn",
        "-ac",
        "1",
        "-ar",
        "16000",
        str(audio_path),
    ]
    process = subprocess.run(command, capture_output=True, text=True)
    if process.returncode != 0:
        raise RuntimeError(process.stderr.strip() or "ffmpeg 提取音频失败")


def transcribe_audio(
    model: WhisperModel, audio_path: Path, language: str
) -> tuple[str, str | None]:
    selected_language = None if language == "auto" else language
    segment_iterator, info = model.transcribe(
        str(audio_path),
        language=selected_language,
        vad_filter=True,
        condition_on_previous_text=False,
    )

    text_parts: list[str] = []
    for segment in segment_iterator:
        text = segment.text.strip()
        if not text:
            continue
        text_parts.append(text)

    combined_text = " ".join(text_parts).strip()
    detected_language = getattr(info, "language", None)
    return combined_text, detected_language


def write_results_csv(output_path: Path, records: Iterable[TaskResult]) -> None:
    field_names = list(TaskResult.__annotations__.keys())
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=field_names)
        writer.writeheader()
        for record in records:
            writer.writerow(asdict(record))


def write_merged_excel(output_path: Path, records: Iterable[TaskResult]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "scripts"
    sheet.append(["链接", "脚本"])
    for record in records:
        if record.status == "ok":
            sheet.append([record.url, record.transcript_text or ""])
        else:
            sheet.append([record.url, ""])
    workbook.save(output_path)


def write_merged_txt(output_path: Path, records: Iterable[TaskResult]) -> None:
    with output_path.open("w", encoding="utf-8") as handle:
        for record in records:
            transcript = (record.transcript_text or "").replace("\n", " ").strip()
            handle.write(f"{record.url}\t{transcript}\n")


def process_urls(
    urls: list[str],
    output_dir: Path,
    model: WhisperModel,
    ffmpeg_bin: str,
    language: str,
    timeout: int,
    keep_audio: bool,
) -> list[TaskResult]:
    directories = ensure_directories(output_dir)
    used_names: set[str] = set()
    records: list[TaskResult] = []

    iterator = tqdm(urls, desc="转写进度") if tqdm else urls

    for index, url in enumerate(iterator, start=1):
        started_at = time.time()
        name = build_base_name(url, index, used_names)
        audio_path = directories["audio"] / f"{name}.wav"

        try:
            extract_audio_from_url(ffmpeg_bin, url, audio_path, timeout=timeout)
            transcript_text, detected_language = transcribe_audio(
                model=model,
                audio_path=audio_path,
                language=language,
            )

            records.append(
                TaskResult(
                    index=index,
                    name=name,
                    url=url,
                    status="ok",
                    transcript_text=transcript_text,
                    error=None,
                    elapsed_seconds=round(time.time() - started_at, 3),
                    detected_language=detected_language,
                )
            )
        except Exception as error:
            records.append(
                TaskResult(
                    index=index,
                    name=name,
                    url=url,
                    status="error",
                    transcript_text=None,
                    error=str(error),
                    elapsed_seconds=round(time.time() - started_at, 3),
                    detected_language=None,
                )
            )
        finally:
            if not keep_audio and audio_path.exists():
                audio_path.unlink(missing_ok=True)

    return records


def print_summary(records: list[TaskResult], output_dir: Path) -> None:
    total = len(records)
    success = sum(1 for record in records if record.status == "ok")
    failed = sum(1 for record in records if record.status == "error")
    print(f"总数: {total} | 成功: {success} | 失败: {failed}")
    print(f"结果汇总: {output_dir / 'results.csv'}")
    print(f"Excel汇总: {output_dir / 'scripts.xlsx'}")
    print(f"TXT汇总: {output_dir / 'scripts.txt'}")
    if failed:
        print("失败样例（最多 5 条）:")
        for record in [item for item in records if item.status == "error"][:5]:
            print(f"- #{record.index} {record.url} -> {record.error}")


def main() -> int:
    arguments = parse_arguments()
    input_path = Path(arguments.input)
    output_dir = Path(arguments.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        ffmpeg_bin = resolve_ffmpeg_binary(arguments.ffmpeg_bin)
        urls = load_urls(input_path)
        if not urls:
            print("输入文件中没有识别到可用链接。")
            return 1

        model = WhisperModel(
            arguments.model,
            device=arguments.device,
            compute_type=arguments.compute_type,
        )
        records = process_urls(
            urls=urls,
            output_dir=output_dir,
            model=model,
            ffmpeg_bin=ffmpeg_bin,
            language=arguments.language,
            timeout=arguments.timeout,
            keep_audio=arguments.keep_audio,
        )
        summary_path = output_dir / "results.csv"
        merged_excel_path = output_dir / "scripts.xlsx"
        merged_txt_path = output_dir / "scripts.txt"
        write_results_csv(summary_path, records)
        write_merged_excel(merged_excel_path, records)
        write_merged_txt(merged_txt_path, records)
        print_summary(records, output_dir)
        return 0 if any(record.status == "ok" for record in records) else 2
    except Exception as error:
        print(f"运行失败: {error}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
